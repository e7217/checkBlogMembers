#-*- coding:utf-8 -*-

from selenium import webdriver
from webdriver_manager.chrome import ChromeDriverManager
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.support.ui import WebDriverWait

import time, datetime
import openpyxl
from openpyxl.styles import Border, Side, PatternFill

startTime = datetime.datetime.now()

options = Options()
options.page_load_strategy = 'eager'
driver = webdriver.Chrome(ChromeDriverManager().install(), options=options)
wait = WebDriverWait(driver, 20)

# print(driver)
print(
    f'''
    제작자 : 엄창용
    누락 범인 찾기 ver. 1.0
    
     ****** dear. 미향 센세 ******
    |                          
    | 냠냠 어디갈까요           
    | 
     ***************  from. 엄창용        
                     
    '''
)
driver.get('https://www.naver.com')
for i in range(0,5):
    print(f'{5-i}초 후 프로그램을 시작합니다.')
    time.sleep(1)

# 포스트별 공감유저의 리스트 추출
def checkMembers(post, curMembers):
    # 포스트 접근
    driver.get(post["postUrl"])
    # 인플루언서 사이트일 경우, #document 아래 접근을 위해 다음과 같은 조치가 필요.
    if 'in.naver.com' in post["postUrl"]:
        wait.until(lambda d: d.find_element_by_tag_name("iframe"))
        driver.switch_to.frame(driver.find_elements_by_tag_name("iframe")[0])

    driver.execute_script("window.scrollTo(0, document.body.scrollHeight);")

    # 공감버튼 탐색
    wait.until(lambda d: d.find_element_by_class_name('btn_like_more'))
    # 아래로 스크롤링
    driver.execute_script("window.scrollTo(0, document.body.scrollHeight);")
    # 공감유저수 링크
    driver.find_element_by_class_name('btn_like_more').send_keys(Keys.ENTER)

    # driver.find_element_by_xpath('//*[@id="ct"]/div[4]/div[3]/div/div[1]/a').send_keys(Keys.ENTER)

    # 공감 유저수 페이지 접근 체크
    wait.until(lambda d: d.find_elements_by_xpath('//*[@id="ct"]/div[4]/div/ul/li'))

    # 가장 아래까지 스크롤 다운
    last_height=0
    while True:
        driver.execute_script("window.scrollTo(0, document.body.scrollHeight);")

        # 1초 대기
        time.sleep(1)

        # 스크롤 다운 후 스크롤 높이 다시 가져옴
        new_height = driver.execute_script("return document.body.scrollHeight")
        if new_height == last_height:
            break
        last_height = new_height

    # 공감유저수 추출
    members = driver.find_elements_by_xpath('//*[@id="ct"]/div[4]/div/ul/li')
    memberList = []
    for i in members:
        member = i.text.split('\n')[0]
        # print(member)
        # 비교를 위한 소문자화
        memberList.append(member.lower())

    # 비교를 위한 소문자화
    for i in range(0, len(curMembers)):
        curMembers[i] = curMembers[i].lower()
    print('curMembers: ',curMembers)
    # 공감유저 리스트와 확인대상 리스트 대조
    setMemberList = set(memberList)
    setCurMembers = set(curMembers)
    # print('setMemberList : ', setMemberList)
    # print('setCurMembers : ', setCurMembers)
    # 교집합
    intersection = set(memberList) & set(curMembers)
    print('공통 유저 : ', intersection)
    # 차집합
    missUsers = set(curMembers) - intersection
    print('누락 유저 : ', missUsers)
    # 포스트 당사자는 제외
    if post["postUser"] in missUsers:
        missUsers.remove(post["postUser"])
    # 확인대상 유저 리스트 중 공감하지 않은 유저 리스트 추출
    return post, missUsers

# 엑셀 시트에서 공감포스트 및 확인대상 유저리스트 불러오기
wbLoad = openpyxl.load_workbook('targetList.xlsx')
datas = wbLoad["Sheet1"]
# print(datas)
posts = []
members = []
for i in datas:
    # 셀 내용이 칼럼 내용일 경우 패스
    if i[0].value == '포스트유저' or i[1].value == '포스트 주소' or i[1].value == '검토멤버' :
        pass
    else :
        try :
            # 모바일 주소가 아닌 경우, 모바일 주소로 변환
            i[1].value = (i[1].value).replace('/blog.naver.com/', '/m.blog.naver.com/')
            if 'naver' in i[1].value:
                posts.append({"postUser" : i[0].value, "postUrl" : i[1].value})
        except:
            pass
        if i[2].value is not None:
            members.append(i[2].value)
        # print("x : ", x)
        # print("value : ", i[1].value)
# 포스트 주소들 출력
# print(posts)
# 확인대상 맴버리스트 출력
# print(members)

# 유저기준 정리
finalChart = {}
# 포스트 기준 정리
finalChart2 = []
cnt = 1
for post in posts:
    try:
        targetPost, missUsers = checkMembers(post, members)
    except Exception as e:
        print(e)
        targetPost = post
        missUsers = ['pass']
        pass
    # 포스트별 공감 미수행 유저 출력
    print(
        f'''
    processing : {cnt}/{len(posts)}
    targetPost : {targetPost}
    missingMember(s) : {missUsers}
    ''')
    # 딕셔너리화
    for i in missUsers:

        if i in finalChart.keys():
            finalChart[i]["cnt"] = finalChart[i]["cnt"]+1
            finalChart[i]["posts"].append(targetPost["postUrl"])
        else :
            finalChart[i] = {"cnt":1, "posts":[targetPost["postUrl"]] }
    finalChart2.append({ "post" : post, "missUsers": missUsers })
    cnt += 1
    # print(finalChart)

# 엑셀에서 열크기 자동 맞추기 함수
def AutoFitColumnSize(worksheet, columns=None, margin=2):
    for i, column_cells in enumerate(worksheet.columns):
        is_ok = False
        if columns == None:
            is_ok = True
        elif isinstance(columns, list) and i in columns:
            is_ok = True

        if is_ok:
            length = max(len(str(cell.value)) for cell in column_cells)
            worksheet.column_dimensions[column_cells[0].column_letter].width = length + margin

    return worksheet

# 엑셀 저장용 세트 생성
wbSave = openpyxl.Workbook()
# 스타일 관련 - 셀 테두리 세트
thin = Side(border_style="thin", color="000000")
medium = Side(border_style="medium", color="000000")
double = Side(border_style="double", color="000000")
# 활성화된 시트 선택
wbSaveSheet = wbSave.active
# 기준점 용도
rowCnt = 2
# 칼럼 생성
wbSaveSheet[f"b{rowCnt}"] = "멤버이름"
wbSaveSheet[f"b{rowCnt}"].border = Border(top=double, left=double, right=thin, bottom=double)
wbSaveSheet[f"c{rowCnt}"] = "누락수"
wbSaveSheet[f"c{rowCnt}"].border = Border(top=double, left=thin, right=thin, bottom=double)
wbSaveSheet[f"d{rowCnt}"] = "포스트주소"
wbSaveSheet[f"d{rowCnt}"].border = Border(top=double, left=thin, right=double, bottom=double)
# 항목 생성
for member in finalChart:
    print('processing...')
    wbSaveSheet[f"b{rowCnt+1}"] = member
    wbSaveSheet[f"b{rowCnt+1}"].fill = PatternFill("solid", fgColor="00FFFF00")
    wbSaveSheet[f"b{rowCnt+1}"].border = Border(top=thin, left=double, right=thin, bottom=thin)
    wbSaveSheet[f"c{rowCnt+1}"] = finalChart[member]["cnt"]
    wbSaveSheet[f"c{rowCnt+1}"].border = Border(top=thin, left=thin, right=thin, bottom=thin)
    wbSaveSheet[f"d{rowCnt+1}"] = str(finalChart[member]["posts"])
    wbSaveSheet[f"d{rowCnt+1}"].border = Border(top=thin, left=thin, right=double, bottom=thin)
    rowCnt += 1
    print(
        f'''
        member : {member}
        Missing : {str(finalChart[member]["cnt"])}
        posts : {str(finalChart[member]["posts"])}
        ''')
# 최하단 셀 테두리
wbSaveSheet[f"b{rowCnt+1}"].border = Border(top=double)
wbSaveSheet[f"c{rowCnt+1}"].border = Border(top=double)
wbSaveSheet[f"d{rowCnt+1}"].border = Border(top=double)
wbSaveSheet = AutoFitColumnSize(wbSaveSheet, margin=4)

# 포스트 기준 시트 생성
wbSaveSheet2 = wbSave.create_sheet('포스트 기준')
rowCnt=2
for i in finalChart2:
    if rowCnt == 2:
        wbSaveSheet2.cell(row=rowCnt, column=2, value='포스트 유저')
        wbSaveSheet2.cell(row=rowCnt, column=2).border = Border(top=double, left=double, right=thin, bottom=double)
        wbSaveSheet2.cell(row=rowCnt, column=3, value='포스트 주소')
        wbSaveSheet2.cell(row=rowCnt, column=3).border = Border(top=double, left=thin, right=thin, bottom=double)
        wbSaveSheet2.cell(row=rowCnt, column=4, value='누락 멤버')
        wbSaveSheet2.cell(row=rowCnt, column=4).border = Border(top=double, left=thin, right=double, bottom=double)
    elif rowCnt > 2:
        wbSaveSheet2.cell(row=rowCnt, column=2, value=i['post']["postUser"])
        wbSaveSheet2.cell(row=rowCnt, column=2).border = Border(top=thin, left=double, right=thin, bottom=thin)
        wbSaveSheet2.cell(row=rowCnt, column=3, value=i['post']["postUrl"])
        wbSaveSheet2.cell(row=rowCnt, column=3).border = Border(top=thin, left=thin, right=thin, bottom=thin)
        wbSaveSheet2.cell(row=rowCnt, column=4, value=str(list(i['missUsers'])))
        wbSaveSheet2.cell(row=rowCnt, column=4).border = Border(top=thin, left=thin, right=double, bottom=thin)
        wbSaveSheet2.cell(row=rowCnt, column=4).fill = PatternFill("solid", fgColor="00FFFF00")
    rowCnt += 1
wbSaveSheet2.cell(row=rowCnt, column=2).border = Border(top=double)
wbSaveSheet2.cell(row=rowCnt, column=3).border = Border(top=double)
wbSaveSheet2.cell(row=rowCnt, column=4).border = Border(top=double)
wbSaveSheet2 = AutoFitColumnSize(wbSaveSheet2, margin=6)

endTime = datetime.datetime.now()
print(
    f'''
    {'='*30}
    - 대상 포스트의 수 : {len(posts)}
    - 검토대상유저의 수 : {len(members)} 
    - 소요시간 : {endTime - startTime}
    {'='*30}
    
    complete!!
    '''
)
#
# 엑셀 파일 생성
wbSave.save('results.xlsx')
for i in range(0,10):
    print(f'{10-i}초 후 종료됩니다.')
    time.sleep(1)
driver.quit()

